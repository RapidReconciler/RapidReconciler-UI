#!/usr/bin/env python3
"""Generate the schema fixtures that run-fingerprint-tests.mjs needs.

WHY THIS EXISTS (UI-176). `fingerprints.json` names 11 fixtures under
`fixtures/`, and that directory is gitignored at `.gitignore:20` -- deliberately,
because real customer F4095 extracts and Transaction Detail exports get staged
there during analysis work. So CI checked out a manifest naming 11 files that do
not exist, and `analyzer-tests.yml` failed every run from at least 2026-06-08 to
2026-09-02. A permanently red suite is not a gate; it is a badge nobody reads.

Reverting the gitignore was not an option -- it is the thing keeping customer
data out of a public repo. This generates the fixtures instead, from literals in
a tracked script, so no customer bytes can enter by construction.

WHAT IT EMITS, AND WHAT THAT DOES NOT COVER. The suite is schema-only: for each
fixture it finds a row in rows 1-3 of any sheet containing one of the manifest's
anchor tokens, then asserts every `requiredHeaders` entry is present after
normalizing (lowercase, whitespace stripped). It never reads a data row and it
never calls the analyzer's `detect()`. So these fixtures carry a title row and a
header row and NOTHING ELSE.

That is deliberate rather than lazy. Emitting invented data rows would imply the
suite exercises parsing or classification, and it does not -- and a fixture that
looks like a sample export invites someone to trust its numbers. If behaviour
coverage is wanted, that is a different suite with different fixtures, and it
should be built knowing this one does not provide it.

THE HEADER TEXT IS THE MANIFEST'S OWN NORMALIZED TOKEN, e.g. `branch/plant`
rather than `Branch/Plant`. Normalization is idempotent, so it satisfies the
same assertion real header text does. Nothing here guesses at a real export's
capitalization or spacing -- that would be inventing an identifier, and the
suite has never checked it anyway.

THE MANIFEST IS THE SINGLE SOURCE. Headers are read from `fingerprints.json`,
never duplicated here, so this script cannot drift from what the suite asserts.
Add a fixture to the manifest and it is generated on the next run.

IT WILL NOT OVERWRITE AN EXISTING FIXTURE unless you pass --force. On a
developer box that directory holds real extracts; clobbering one would destroy
data the gitignore exists to protect. In CI the directory is empty, so every
fixture is created.

Usage:
    python make-fixtures.py            # create only what is missing
    python make-fixtures.py --force    # regenerate everything (DESTRUCTIVE)
    python make-fixtures.py --check    # verify fixtures satisfy the manifest
"""

import io
import json
import os
import sys

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    raise SystemExit(2)

HERE = os.path.dirname(os.path.abspath(__file__))
MANIFEST = os.path.join(HERE, "fingerprints.json")


def normalize(v):
    """Mirror Helpers.readHeaders / the suite's own normalize()."""
    if v is None:
        return ""
    return "".join(str(v).split()).lower()


def load_manifest():
    with io.open(MANIFEST, encoding="utf-8") as fh:
        return json.load(fh)


def header_row(fx):
    """Required headers, plus any anchor token not already among them.

    The suite matches the header row by anchor token FIRST, then checks the
    required list. A manifest whose anchors are not all in requiredHeaders is
    legitimate -- `system-status` is one -- so the anchors have to be added or
    the row would never be found.
    """
    headers = [normalize(h) for h in fx["requiredHeaders"]]
    for a in fx["anchorTokens"]:
        n = normalize(a)
        if n not in headers:
            headers.append(n)
    return headers


def write_fixture(path, fx):
    wb = Workbook()
    ws = wb.active
    # Sheet name: the template it stands for, trimmed to Excel's 31-char cap
    # and stripped of the characters Excel refuses in a sheet name.
    name = fx["template"]
    for ch in "[]:*?/\\":
        name = name.replace(ch, " ")
    ws.title = name.strip()[:31] or "Sheet1"
    # Row 1 is a title, matching how the real exports lead. The suite scans
    # rows 1-3, so the header row sitting at row 2 is exercised rather than
    # the trivial row-1 case.
    ws.append(["Generated schema fixture for the %s template - headers only, no data"
               % fx["template"]])
    ws.append(header_row(fx))
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def check_fixture(path, fx):
    """Replicate the suite's assertion. Returns (ok, message).

    This exists because the suite itself is Node, and there is no Node runtime
    on the box this was written on -- so the only local proof available is to
    re-implement the same check. It is NOT a substitute for the real run; CI is
    the authority. It catches a broken generator in one second instead of one
    push.
    """
    if not os.path.exists(path):
        return False, "missing"
    wb = load_workbook(path, data_only=True)
    anchors = [normalize(a) for a in fx["anchorTokens"]]
    required = [normalize(h) for h in fx["requiredHeaders"]]
    try:
        for ws in wb.worksheets:
            for r in range(1, 4):
                row = [normalize(c.value) for c in ws[r] if c.value is not None]
                if any(a in row for a in anchors):
                    missing = [h for h in required if h not in row]
                    if missing:
                        return False, 'sheet "%s" row %d missing: %s' % (
                            ws.title, r, ", ".join(missing))
                    return True, 'sheet "%s" row %d, all %d headers' % (
                        ws.title, r, len(required))
        return False, "no row in rows 1-3 of any sheet holds an anchor token (%s)" % (
            ", ".join(anchors))
    finally:
        wb.close()


def main(argv):
    force = "--force" in argv
    check_only = "--check" in argv

    manifest = load_manifest()
    fixtures = manifest["fixtures"]
    created = skipped = 0
    failures = []

    for fx in fixtures:
        path = os.path.join(HERE, fx["file"])
        rel = fx["file"]

        if not check_only:
            if os.path.exists(path) and not force:
                print("  skip    %-40s (exists - use --force to regenerate)" % rel)
                skipped += 1
            else:
                write_fixture(path, fx)
                print("  write   %-40s" % rel)
                created += 1

        ok, msg = check_fixture(path, fx)
        print("  %-7s %-40s %s" % ("PASS" if ok else "FAIL", rel, msg))
        if not ok:
            failures.append((rel, msg))

    print()
    if not check_only:
        print("%d written, %d skipped." % (created, skipped))
    print("%d of %d fixtures satisfy the manifest." % (len(fixtures) - len(failures),
                                                       len(fixtures)))
    if failures:
        print()
        print("FAILURES:")
        for rel, msg in failures:
            print("  %-40s %s" % (rel, msg))
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
