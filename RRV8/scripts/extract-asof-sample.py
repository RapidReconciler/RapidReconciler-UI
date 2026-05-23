#!/usr/bin/env python3
"""
extract-asof-sample.py — read a legacy As-Of xlsx export and emit
`RRV8/data/as-of.json` for V8 demo mode.

The legacy AngularJS As Of page exports a flat workbook:
  Row 1 : banner ("As Of Generated <ts>")
  Row 2 : column headers
  Row 3 : Grand Summaries row (totals for the money/qty columns)
  Row 4+: data rows

Usage:
    python3 extract-asof-sample.py \
        --xlsx "<path to AsOf_<period>_<stamp>.xlsx>" \
        --period 2016-08-27 \
        --out RRV8/data/as-of.json
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl required: python3 -m pip install --user openpyxl\n")
    sys.exit(2)


def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx",   required=True, type=Path)
    ap.add_argument("--period", required=True, help="As-Of date, e.g. 2016-08-27")
    ap.add_argument(
        "--out",
        type=Path,
        help="Output JSON path. Defaults to <repo>/RRV8/data/as-of.json",
    )
    args = ap.parse_args()

    out_path = args.out or (
        Path(__file__).resolve().parent.parent / "data" / "as-of.json"
    )

    if not args.xlsx.exists():
        raise SystemExit(f"Workbook not found: {args.xlsx}")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb.active

    banner = _txt(ws.cell(1, 1).value)
    headers = [_txt(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)]
    # Row 3 = "Grand Summaries" totals row, captured separately.
    totals_row = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]

    rows = []
    for r in range(4, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if first is None or first == "":
            continue
        rec = {}
        for c, h in enumerate(headers, start=1):
            if not h:
                continue
            v = ws.cell(r, c).value
            if v is None:
                continue
            # Trim trailing spaces from string fields (the legacy export
            # right-pads many string columns).
            if isinstance(v, str):
                v = v.strip()
                if not v:
                    continue
            rec[h] = v
        rows.append(rec)

    # Sum totals from the data for fields the analyst cares about.
    # Mirrors the legacy "Filtered Totals" card.
    sum_keys = ["Quantity", "Amount", "QtyVar", "AmtVar"]
    totals = {k: 0.0 for k in sum_keys}
    for rec in rows:
        for k in sum_keys:
            v = rec.get(k)
            if isinstance(v, (int, float)):
                totals[k] += float(v)

    # Round to 2 decimals where appropriate to keep the JSON readable.
    totals = {k: round(v, 4) for k, v in totals.items()}

    payload = {
        "_meta": {
            "banner":       banner,
            "period":       args.period,
            "exportedAt":   datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "columns":      [h for h in headers if h],
            "rowCount":     len(rows),
        },
        "totals": totals,
        "rows":   rows,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    print(
        f"Wrote {out_path} — period={args.period}, "
        f"{len(rows):,} rows, {len(payload['_meta']['columns'])} columns"
    )


if __name__ == "__main__":
    main()
